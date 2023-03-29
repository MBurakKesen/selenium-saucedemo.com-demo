from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
from constants import sabitler as sabit




class Test_DemoClass():
    def setup_method(self):
 
        self.driver=webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(sabit.Url)
        self.folderPath=str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)


    def getData():
        excelFile=openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet=excelFile["Sheet1"]

        totalRows=selectedSheet.max_row
        data=[]
        for i in range(2,totalRows+1):
            username=selectedSheet.cell(i,1).value
            password=selectedSheet.cell(i,2).value
            tupleData=(username,password)
            data.append(tupleData)
        return data
    
    


    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput=self.driver.find_element(By.ID,"user-name")

        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput=self.driver.find_element(By.ID,"password")
        
        passwordInput.send_keys(password)
        usernameInput.send_keys(username)
        loginBtn=self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage=True
        assert errorMessage==True

    def waitForElementVisible(self, locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_all_elements_located(locator))